import os
import csv
import glob
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter, coordinate_to_tuple
from openpyxl.styles import numbers
import subprocess

def csv_to_xlsx(csv_folder):
    # Get the list of CSV files in the folder
    csv_files = glob.glob(csv_folder + '/*.csv')

    # Iterate over each CSV file
    for csv_file in csv_files:
        # Extract the filename without the extension
        file_name = os.path.splitext(os.path.basename(csv_file))[0]

        # Create a new workbook and sheet
        workbook = Workbook()
        sheet = workbook.active

        # Read the CSV file using csv module with delimiter auto-detection
        with open(csv_file, 'r') as file:
            dialect = csv.Sniffer().sniff(file.read(8192))
            file.seek(0)
            reader = csv.reader(file, dialect)
            data = list(reader)

        # Convert numeric values to floats
        for row in data:
            for i, cell in enumerate(row):
                try:
                    row[i] = float(cell)
                except ValueError:
                    pass

        # Write the data to the sheet
        for row in data:
            sheet.append(row)

        # Apply number format to cells in numeric columns
        for column in sheet.columns:
            is_numeric = all(isinstance(cell.value, (int, float)) for cell in column if cell.value is not None)
            if is_numeric:
                column_letter = column[0].column_letter
                for cell in column:
                    cell.number_format = numbers.FORMAT_NUMBER_00

        # Save the workbook
        xlsx_file = os.path.join(csv_folder, file_name + '.xlsx')
        workbook.save(xlsx_file)


def xlsx_merge_copy(folder_path, existing_workbook, target_sheet, output_file, column_indices, paste_cell):
    # Convert the top-left cell address to coordinates
    paste_row, paste_column = coordinate_to_tuple(paste_cell)

    # Load the existing workbook
    workbook = load_workbook(existing_workbook)

    # Get the target sheet from the existing workbook
    sheet = workbook[target_sheet]

    # Create a new workbook to store the data from XLSX files
    new_workbook = Workbook()

    # Remove the default sheet created in the new workbook
    new_workbook.remove(new_workbook.active)

    # Iterate over the files in the folder
    files = os.listdir(folder_path)

    for file_name in files:
        if file_name.endswith(".xlsx"):  # Check if the file is an XLSX file
            # Create a new sheet in the new workbook for each file
            new_sheet = new_workbook.create_sheet()

            # Copy the target sheet to the new sheet
            for row in sheet.iter_rows(min_row=1, values_only=True):
                new_sheet.append(row)

            file_path = os.path.join(folder_path, file_name)
            source_workbook = load_workbook(file_path, read_only=True)

            # Assuming the data is in the first sheet of the source workbook
            source_sheet = source_workbook.active

            # Copy the data from the source sheet to the new sheet
            for row_index, row in enumerate(source_sheet.iter_rows(values_only=True), start=1):  # Start from row 1
                if row_index > 18:  # Skip the first 18 rows
                    for column_index, index in enumerate(column_indices, start=1):
                        if index < len(row):
                            cell_value = row[index]
                            if cell_value is not None:
                                try:
                                    cell_value = float(cell_value)
                                except ValueError:
                                    pass
                            new_row_index = paste_row + row_index - 19
                            new_column_index = paste_column + column_index - 1
                            new_sheet.cell(row=new_row_index, column=new_column_index, value=cell_value)

            # Rename the new sheet based on the last 30 characters of the XLSX file name
            sheet_name = os.path.splitext(file_name)[0][-27:]
            new_sheet.title = sheet_name

            print(f"Finished processing {file_name}")

    # Rename the new workbook if a file with the same name already exists
    counter = 1
    new_file_path = os.path.join(folder_path, output_file)
    while os.path.exists(new_file_path):
        new_file_name = f"ProcessedData_{counter}.xlsx"
        new_file_path = os.path.join(folder_path, new_file_name)
        counter += 1

    # Save the new workbook
    new_workbook.save(new_file_path)

    # Open the new file
    subprocess.call(["open", new_file_path])


# Example usage

# target folder for csv files 
csv_folder = '/Users/akin.o.akinjogbin/Desktop/July 2023 Tine Testing Raw data/Pull Test/Batch 2 Test Data (12mm Pocket)/Presentation'

#target folder to dump processes csv file 
xlsx_folder = csv_folder

#details of excel sheet template to use 
existing_workbook = "/Users/akin.o.akinjogbin/Desktop/Python Repository/Working Codes/Analysissht.xlsx"
target_sheet = "Data Sheet"

#output excel sheet  details 
output_file = "ProcessedData.xlsx" 

#Adjust columss to copy
column_indices = [1, 2, 3]     
paste_cell = "A2"

csv_to_xlsx(csv_folder)
xlsx_merge_copy(xlsx_folder, existing_workbook, target_sheet, output_file, column_indices, paste_cell)
